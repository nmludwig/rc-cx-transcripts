#!/usr/bin/env python3
"""
RingCentral RingCX Transcript Downloader — Web Server
Uses the RingCX Integration API (cx/integration/v1) with 3-legged OAuth.

Auth is a TWO-STEP process:
  Step 1: Standard RingCentral OAuth → RC access token (ReadAccounts scope)
  Step 2: Exchange RC token for a RingCX-specific token via:
          POST https://ringcx.ringcentral.com/api/auth/login/rc/accesstoken
  The RingCX token expires in ~5 minutes and must be refreshed separately.
  Sub-account ID is extracted from the CX token exchange response (agentDetails.accountId).

Key differences from the RingEX/RingSense version:
  - Call discovery:  POST .../interaction-metadata  (returns dialogId + segmentId)
  - Transcript:      GET  .../transcripts/dialogs/{dialogId}/segments/{segmentId}
  - Auth scope:      ReadAccounts only (platform permissions managed in RingCX Admin)
  - Sub-account:     Extracted from CX token exchange response — no separate lookup needed
  - User must be:    RingCX admin with WEM access enabled on their user profile
"""

import sys, os, re, time, threading, uuid, secrets
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlencode
from flask import (Flask, render_template, request, jsonify,
                   send_file, session, redirect, url_for)

try:
    from flask_cors import CORS
except ImportError:
    CORS = None

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Config — override via environment variables in production
# ---------------------------------------------------------------------------
RC_CLIENT_ID     = os.environ.get("RC_CLIENT_ID", "")
RC_CLIENT_SECRET = os.environ.get("RC_CLIENT_SECRET", "")
RC_REDIRECT_URI  = os.environ.get("RC_REDIRECT_URI", "http://localhost:5000/oauth/callback")

RC_AUTH_URL  = "https://platform.ringcentral.com/restapi/oauth/authorize"
RC_TOKEN_URL = "https://platform.ringcentral.com/restapi/oauth/token"

# RingCX token exchange — Step 2 of auth (for CX admin APIs)
CX_TOKEN_URL         = "https://ringcx.ringcentral.com/api/auth/login/rc/accesstoken"
CX_TOKEN_REFRESH_URL = "https://ringcx.ringcentral.com/api/auth/token/refresh"

# RingCX integration base — uses platform.ringcentral.com per RC docs
CX_BASE     = "https://ringcx.ringcentral.com/voice/api"
PLATFORM_BASE = "https://platform.ringcentral.com"

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET", "rc-cx-demo-secret-change-in-prod")
if CORS:
    CORS(app)

OUTPUT_DIR = Path(__file__).parent / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)

jobs = {}   # in-memory job store


# ---------------------------------------------------------------------------
# Routes — auth
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    if session.get("cx_token"):
        return render_template("index.html", authed=True,
                               display_name=session.get("rc_display_name", ""),
                               sub_accounts=session.get("rc_sub_accounts", []))
    return render_template("index.html", authed=False)


@app.route("/login")
def login():
    state = secrets.token_urlsafe(16)
    session["oauth_state"] = state
    params = {
        "response_type": "code",
        "client_id":     RC_CLIENT_ID,
        "redirect_uri":  RC_REDIRECT_URI,
        "state":         state,
        # ReadAccounts is the only OAuth scope that affects RingCX APIs
        "scope":         "ReadAccounts",
    }
    return redirect(RC_AUTH_URL + "?" + urlencode(params))


@app.route("/oauth/callback")
def oauth_callback():
    error = request.args.get("error")
    if error:
        return render_template("error.html", message=f"Login failed: {error}"), 400

    code  = request.args.get("code")
    state = request.args.get("state")
    if not code:
        return render_template("error.html", message="No auth code received."), 400
    if state != session.get("oauth_state"):
        return render_template("error.html", message="Invalid state parameter."), 400

    # Exchange code for tokens
    try:
        resp = requests.post(
            RC_TOKEN_URL,
            auth=(RC_CLIENT_ID, RC_CLIENT_SECRET),
            data={"grant_type": "authorization_code", "code": code,
                  "redirect_uri": RC_REDIRECT_URI},
            timeout=15)
        resp.raise_for_status()
    except requests.exceptions.HTTPError as e:
        code_n = e.response.status_code if e.response else "?"
        return render_template("error.html", message=f"Token exchange failed (HTTP {code_n})."), 400
    except Exception as e:
        return render_template("error.html", message=f"Connection error: {e}"), 503

    rc_data         = resp.json()
    rc_token        = rc_data["access_token"]
    rc_refresh      = rc_data.get("refresh_token", "")

    # ------------------------------------------------------------------
    # Step 2 — Exchange RC token for a RingCX-specific token
    # The CX token is what actually authorizes all CX API calls.
    # It expires in ~5 minutes and carries accountId + agentDetails.
    # ------------------------------------------------------------------
    cx_token, cx_refresh, rc_account_id, sub_accounts, display_name = \
        _exchange_for_cx_token(rc_token)

    if not cx_token:
        return render_template("error.html",
            message="Logged in to RingCentral but could not exchange for a RingCX token. "
                    "Make sure this user is a RingCX admin with WEM access enabled."), 403

    session["rc_token"]         = rc_token        # keep for re-exchange if needed
    session["rc_refresh_token"] = rc_refresh
    session["cx_token"]         = cx_token        # used for all CX API calls
    session["cx_refresh_token"] = cx_refresh
    session["cx_token_time"]    = datetime.now().timestamp()
    session["rc_account_id"]    = rc_account_id
    session["rc_display_name"]  = display_name
    session["rc_sub_accounts"]  = sub_accounts
    session.pop("oauth_state", None)
    return redirect(url_for("index"))


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("index"))


@app.route("/debug/token")
def debug_token():
    """Temporary debug route — remove before production."""
    if not session.get("cx_token"):
        return jsonify({"error": "not logged in"}), 401
    return jsonify({
        "cx_token_full": session.get("cx_token", ""),
        "rc_token_full": session.get("rc_token", ""),
        "rc_account_id": session.get("rc_account_id"),
        "sub_accounts":  session.get("rc_sub_accounts"),
    })


# ---------------------------------------------------------------------------
# CX Token Exchange (Step 2 of auth)
# ---------------------------------------------------------------------------

def _exchange_for_cx_token(rc_token: str):
    """
    Exchange a RingCentral OAuth token for a RingCX-specific token.
    Returns (cx_token, cx_refresh, rc_account_id, sub_accounts, display_name)
    or (None, None, None, [], "") on failure.
    """
    try:
        resp = requests.post(
            CX_TOKEN_URL + "?includeRefresh=true",
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            data={"rcAccessToken": rc_token, "rcTokenType": "Bearer"},
            timeout=15)

        if resp.status_code != 200:
            return None, None, None, [], ""

        data = resp.json()
        cx_token   = data.get("accessToken", "")
        cx_refresh = data.get("refreshToken", "")

        # Pull sub-accounts and display name from agentDetails
        agent_details = data.get("agentDetails", [])
        sub_accounts  = []
        display_name  = ""
        rc_account_id = ""

        for agent in agent_details:
            acct_id   = str(agent.get("accountId", ""))
            acct_name = agent.get("accountName", "")
            first     = agent.get("firstName", "")
            last      = agent.get("lastName", "")
            if not display_name and (first or last):
                display_name = f"{first} {last}".strip()
            if acct_id and not any(s["id"] == acct_id for s in sub_accounts):
                sub_accounts.append({"id": acct_id, "name": acct_name})
            if not rc_account_id and acct_id:
                rc_account_id = acct_id

        return cx_token, cx_refresh, rc_account_id, sub_accounts, display_name

    except Exception as e:
        return None, None, None, [], ""


def _refresh_cx_token(cx_refresh: str):
    """Refresh an expired RingCX token. Returns (new_cx_token, new_cx_refresh)."""
    try:
        resp = requests.post(
            CX_TOKEN_REFRESH_URL,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            data={"refresh_token": cx_refresh, "rcTokenType": "Bearer"},
            timeout=15)
        if resp.status_code == 200:
            data = resp.json()
            return data.get("accessToken", ""), data.get("refreshToken", cx_refresh)
    except Exception:
        pass
    return None, cx_refresh


# ---------------------------------------------------------------------------
# Job API
# ---------------------------------------------------------------------------

@app.route("/api/sub-accounts")
def api_sub_accounts():
    if not session.get("cx_token"):
        return jsonify({"error": "Not authenticated."}), 401
    return jsonify(session.get("rc_sub_accounts", []))


@app.route("/api/start-job", methods=["POST"])
def api_start_job():
    if not session.get("cx_token"):
        return jsonify({"error": "Not authenticated."}), 401

    data          = request.json or {}
    customer_name = data.get("customer_name", "Customer").strip() or "Customer"
    date_from     = data.get("date_from", "")
    date_to       = data.get("date_to", "")
    sub_account_id = data.get("sub_account_id", "")

    if not date_from or not date_to:
        return jsonify({"error": "date_from and date_to are required."}), 400
    if not sub_account_id:
        return jsonify({"error": "sub_account_id is required."}), 400

    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "running", "progress": 0, "log": [],
                    "records": [], "files": {}, "error": None, "summary": None}

    threading.Thread(
        target=run_cx_download_job,
        args=(job_id,
              session["rc_token"],          # RC token — for platform.ringcentral.com integration APIs
              session["rc_refresh_token"],  # RC refresh token
              session["cx_token"],          # CX token — for ringcx.ringcentral.com admin APIs
              session["cx_refresh_token"],
              session["rc_account_id"],
              sub_account_id,
              customer_name,
              date_from,
              date_to),
        daemon=True).start()

    return jsonify({"job_id": job_id})


@app.route("/api/job/<job_id>")
def api_job_status(job_id):
    job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify({"status": job["status"], "progress": job["progress"],
                    "log": job["log"][-50:], "files": list(job["files"].keys()),
                    "error": job["error"], "summary": job.get("summary")})


@app.route("/api/download/<job_id>/<file_type>")
def api_download(job_id, file_type):
    job = jobs.get(job_id)
    if not job or file_type not in job["files"]:
        return jsonify({"error": "File not found"}), 404
    path = Path(job["files"][file_type])
    if not path.exists():
        return jsonify({"error": "File no longer available"}), 404
    return send_file(str(path), as_attachment=True, download_name=path.name)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def job_log(job_id, msg, level="info"):
    jobs[job_id]["log"].append({"t": datetime.now().strftime("%H:%M:%S"),
                                "msg": msg, "level": level})


def refresh_rc_token(rc_refresh: str):
    """Refresh an expired RC token inside a running job."""
    try:
        resp = requests.post(
            RC_TOKEN_URL,
            auth=(RC_CLIENT_ID, RC_CLIENT_SECRET),
            data={"grant_type": "refresh_token", "refresh_token": rc_refresh},
            timeout=15)
        if resp.status_code == 200:
            d = resp.json()
            return d.get("access_token"), d.get("refresh_token", rc_refresh)
    except Exception:
        pass
    return None, rc_refresh


def _cx_headers(token: str) -> dict:
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}


# ---------------------------------------------------------------------------
# Core download job
# ---------------------------------------------------------------------------

def run_cx_download_job(job_id, rc_token, rc_refresh_token, cx_token, cx_refresh_token,
                        rc_account_id, sub_account_id, customer_name, date_from, date_to):
    job = jobs[job_id]
    # Integration APIs use CX token on ringcx.ringcentral.com
    token = cx_token
    refresh_token = cx_refresh_token
    try:
        job_log(job_id, f"Starting RingCX download for {customer_name}")
        job_log(job_id, f"RC Account: {rc_account_id}  |  Sub-account: {sub_account_id}")
        job_log(job_id, f"Date range: {date_from} → {date_to}")
        job["progress"] = 5

        # ------------------------------------------------------------------
        # Step 1 – Collect interaction metadata in sliding 1-hour windows
        # ------------------------------------------------------------------
        job_log(job_id, "Fetching interaction metadata from RingCX…")

        from datetime import timedelta

        def parse_dt(s):
            return datetime.fromisoformat(s + "T00:00:00").replace(tzinfo=timezone.utc)

        start_dt = parse_dt(date_from)
        end_dt   = parse_dt(date_to).replace(hour=23, minute=59, second=59)

        # API requires data to be at least 15 minutes old to avoid processing errors
        safe_ceiling = datetime.now(timezone.utc) - timedelta(minutes=15)
        if end_dt > safe_ceiling:
            end_dt = safe_ceiling
            job_log(job_id, "End date capped to 15 min ago (API processing window).", "warn")

        window   = timedelta(minutes=30)   # 1800 seconds per Calabrio docs

        all_segments = []   # list of {dialogId, segmentId, metadata…}

        cursor = start_dt
        while cursor < end_dt:
            window_end = min(cursor + window, end_dt)
            ts_from = int(cursor.timestamp() * 1000)
            ts_to   = int(window_end.timestamp() * 1000)

            # v2 endpoint — uses segmentEndTime + timeInterval (seconds) + timeZone
            url = (f"{CX_BASE}/integration/v2/admin/reports"
                   f"/accounts/{sub_account_id}/interactionMetadata")

            # Payload: segmentEndTime = end of window, timeInterval = window size in seconds
            window_seconds = int((window_end - cursor).total_seconds())
            payload = {
                "segmentEndTime": window_end.strftime("%Y-%m-%d %H:%M:%S"),
                "timeInterval":   window_seconds,
                "timeZone":       "US/Eastern",
                "pageSize":       200,
            }

            page_token = None
            while True:
                if page_token:
                    payload["pageToken"] = page_token

                for attempt in range(3):
                    r = requests.post(url, json=payload, headers=_cx_headers(token), timeout=30)
                    if r.status_code == 429:
                        job_log(job_id, "Rate limit hit on metadata — waiting 35 s…", "warn")
                        time.sleep(35)
                        continue
                    if r.status_code == 401:
                        new_tok, refresh_token = _refresh_cx_token(refresh_token)
                        if new_tok:
                            token = new_tok
                            job_log(job_id, "CX token refreshed.", "ok")
                        continue
                    break

                if r.status_code != 200:
                    job_log(job_id, f"Metadata fetch failed ({r.status_code}): {r.text[:300]}", "warn")
                    job_log(job_id, f"URL: {url}", "warn")
                    job_log(job_id, f"Payload: {payload}", "warn")
                    break

                body     = r.json()
                segments = body.get("segments", body.get("records", []))
                all_segments.extend(segments)

                page_token = body.get("nextPageToken") or body.get("paging", {}).get("nextPageToken")
                if not page_token:
                    break

                time.sleep(0.6)   # stay well under 2 calls/min for metadata

            cursor = window_end + timedelta(seconds=1)

        job["progress"] = 20
        job_log(job_id, f"Found {len(all_segments)} interaction segments", "ok")

        if not all_segments:
            job.update({"status": "done", "progress": 100,
                        "summary": {"total": 0, "transcripts": 0}})
            job_log(job_id, "No interactions found in this date range.", "warn")
            return

        # ------------------------------------------------------------------
        # Step 2 – Fetch transcript for each segment
        # ------------------------------------------------------------------
        total          = len(all_segments)
        transcript_records = []
        with_transcripts   = 0

        job_log(job_id, f"Fetching transcripts for {total} segments…")
        job_log(job_id,
                f"Estimated time: ~{round(total * 0.5 / 60, 1)} minutes at ~0.5 s/call.", "warn")

        call_count = 0
        for i, seg in enumerate(all_segments):
            dialog_id  = seg.get("dialogId",  seg.get("dialog_id", ""))
            segment_id = seg.get("segmentId", seg.get("segment_id", ""))

            if not dialog_id or not segment_id:
                continue

            job["progress"] = 20 + int(65 * (i / max(total, 1)))

            transcript_url = (
                f"{CX_BASE}/integration/v2/admin/reports"
                f"/accounts/{sub_account_id}"
                f"/transcripts/dialogs/{dialog_id}/segments/{segment_id}"
            )

            transcript_data = None
            for attempt in range(4):
                try:
                    tr = requests.get(transcript_url, headers=_cx_headers(token), timeout=20)
                    if tr.status_code == 429:
                        wait = 30 * (attempt + 1)
                        job_log(job_id, f"Rate limit — waiting {wait} s…", "warn")
                        time.sleep(wait)
                        continue
                    if tr.status_code == 404:
                        break   # no transcript for this segment
                    if tr.status_code == 401:
                        token, refresh_token = refresh_rc_token(refresh_token)
                        continue
                    if tr.status_code == 200:
                        transcript_data = tr.json()
                        break
                except Exception:
                    time.sleep(2)

            # Build unified lines from CX transcript format
            lines = []
            if transcript_data:
                with_transcripts += 1
                utterances = transcript_data.get("transcript", [])
                for u in utterances:
                    name = u.get("participantName", u.get("participantId", "?"))
                    text = u.get("message", u.get("text", "")).strip()
                    ts_ms = u.get("timestamp", 0)
                    try:
                        secs  = int(ts_ms) // 1000 if ts_ms else 0
                        mm    = str(secs // 60).zfill(2)
                        ss    = str(secs % 60).zfill(2)
                        ts_str = f"[{mm}:{ss}]"
                    except Exception:
                        ts_str = ""
                    lines.append(f"{ts_str} {name}: {text}")

            # Pull metadata fields — field names vary by API version
            start_time    = seg.get("startTime",    seg.get("start_time",    ""))
            duration_sec  = int(seg.get("duration", seg.get("durationMs", 0)) or 0)
            if duration_sec > 10000:   # likely milliseconds
                duration_sec = duration_sec // 1000
            direction     = seg.get("direction",    seg.get("callDirection", ""))
            from_number   = seg.get("fromNumber",   seg.get("ani",           ""))
            from_name     = seg.get("fromName",     seg.get("agentName",     ""))
            to_number     = seg.get("toNumber",     seg.get("dnis",          ""))
            to_name       = seg.get("toName",       seg.get("queueName",     ""))
            queue_name    = seg.get("queueName",    seg.get("skillName",     ""))
            agent_name    = seg.get("agentName",    "")

            transcript_records.append({
                "dialog_id":    dialog_id,
                "segment_id":   segment_id,
                "start_time":   start_time,
                "duration_sec": duration_sec,
                "direction":    direction,
                "from_number":  from_number,
                "from_name":    from_name,
                "to_number":    to_number,
                "to_name":      to_name,
                "queue_name":   queue_name,
                "agent_name":   agent_name,
                "channel":      transcript_data.get("channelClass", "VOICE") if transcript_data else "VOICE",
                "has_transcript": transcript_data is not None,
                "transcript":   "\n".join(lines),
            })

            call_count += 1
            if call_count % 10 == 0:
                job_log(job_id,
                        f"Processed {call_count}/{total} segments "
                        f"({with_transcripts} transcripts so far)")

            # Refresh CX token every 20 calls (CX tokens expire in ~5 minutes)
            if call_count % 20 == 0 and refresh_token:
                new_tok, new_refresh = _refresh_cx_token(refresh_token)
                if new_tok:
                    token = new_tok
                    refresh_token = new_refresh
                    job_log(job_id, "CX access token refreshed.", "ok")

            time.sleep(0.5)

        job_log(job_id, f"{with_transcripts} of {total} segments have transcripts", "ok")
        job["progress"] = 88

        # ------------------------------------------------------------------
        # Step 3 – Build output files
        # ------------------------------------------------------------------
        slug       = re.sub(r"[^a-z0-9]+", "_", customer_name.lower()).strip("_")
        date_stamp = datetime.now().strftime("%Y%m%d")

        job_log(job_id, "Building Excel spreadsheet…")
        xlsx_path = OUTPUT_DIR / f"cx_transcripts_{slug}_{date_stamp}.xlsx"
        build_excel(transcript_records, customer_name, date_from, date_to, xlsx_path)
        job["files"]["xlsx"] = str(xlsx_path)
        job_log(job_id, f"Excel saved: {xlsx_path.name}", "ok")

        job["progress"] = 94

        job_log(job_id, "Building PDF…")
        pdf_path = OUTPUT_DIR / f"cx_transcripts_{slug}_{date_stamp}.pdf"
        build_pdf(transcript_records, customer_name, date_from, date_to, pdf_path)
        job["files"]["pdf"] = str(pdf_path)
        job_log(job_id, f"PDF saved: {pdf_path.name}", "ok")

        job.update({
            "progress": 100, "status": "done",
            "summary": {
                "total":      len(transcript_records),
                "transcripts": with_transcripts,
                "customer":   customer_name,
                "date_from":  date_from,
                "date_to":    date_to,
            }
        })
        job_log(job_id, "All done! Files are ready to download.", "ok")

    except Exception as e:
        import traceback
        job.update({"status": "error", "error": str(e)})
        job_log(job_id, f"Error: {e}", "error")
        job_log(job_id, traceback.format_exc(), "error")


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

def build_excel(records, customer_name, date_from, date_to, out_path):
    wb = openpyxl.Workbook()
    RC_BLUE   = "0066CC"
    RC_ORANGE = "FF6A00"
    DARK      = "1A1A1A"
    WHITE     = "FFFFFF"
    LIGHT     = "F5F5F5"
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"RingCentral RingCX Transcript Report — {customer_name}"
    c.font  = Font(name="Calibri", size=16, bold=True, color=WHITE)
    c.fill  = PatternFill("solid", fgColor=RC_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:H2")
    c2 = ws["A2"]
    c2.value = f"{date_from} to {date_to} | Generated {datetime.now().strftime('%B %d, %Y')} | Confidential"
    c2.font  = Font(name="Calibri", size=10, color="555555")
    c2.fill  = PatternFill("solid", fgColor=LIGHT)
    c2.alignment = Alignment(horizontal="center", vertical="center")

    with_trans  = len([r for r in records if r["has_transcript"]])
    total_sec   = sum(r["duration_sec"] for r in records)
    hrs, rem    = divmod(total_sec, 3600)
    mins        = rem // 60

    ws.append([])
    for row_idx, (label, value) in enumerate([
        ("Total Interactions", str(len(records))),
        ("With Transcripts",   str(with_trans)),
        ("Without Transcripts", str(len(records) - with_trans)),
        ("Total Talk Time",    f"{hrs}h {mins}m"),
        ("Date From",          date_from),
        ("Date To",            date_to),
    ], start=4):
        ws.cell(row=row_idx, column=1).value = label
        ws.cell(row=row_idx, column=1).font  = Font(name="Calibri", size=11, bold=True, color=DARK)
        ws.cell(row=row_idx, column=1).fill  = PatternFill("solid", fgColor=LIGHT)
        ws.cell(row=row_idx, column=2).value = value
        ws.cell(row=row_idx, column=2).font  = Font(name="Calibri", size=11, color=DARK)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22

    # --- All Interactions sheet ---
    ws2 = wb.create_sheet("All Interactions")
    headers = ["Date", "Time", "Direction", "Duration", "From Name", "From Number",
               "To Name", "To Number", "Queue", "Agent", "Channel", "Has Transcript"]
    widths  = [14, 10, 12, 10, 22, 18, 22, 18, 22, 22, 10, 14]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws2.cell(row=1, column=col)
        cell.value = h
        cell.font  = Font(name="Calibri", size=11, bold=True, color=WHITE)
        cell.fill  = PatternFill("solid", fgColor=RC_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[1].height = 24

    for ri, rec in enumerate(records, 2):
        try:
            dt = datetime.fromisoformat(rec["start_time"].replace("Z", "+00:00"))
            ds = dt.strftime("%Y-%m-%d")
            ts = dt.strftime("%I:%M %p")
        except Exception:
            ds = rec["start_time"][:10] if rec["start_time"] else ""
            ts = ""
        dm, ds2 = divmod(int(rec["duration_sec"]), 60)
        for col, val in enumerate([
            ds, ts, rec["direction"], f"{dm}m {ds2}s",
            rec["from_name"], rec["from_number"],
            rec["to_name"],   rec["to_number"],
            rec["queue_name"], rec["agent_name"],
            rec["channel"],
            "Yes" if rec["has_transcript"] else "No",
        ], 1):
            cell = ws2.cell(row=ri, column=col)
            cell.value = val
            cell.font  = Font(name="Calibri", size=10, color=DARK)
            cell.alignment = Alignment(vertical="top")
            cell.border = border
        ws2.row_dimensions[ri].height = 15

    # --- Transcripts sheet ---
    ws3 = wb.create_sheet("Transcripts")
    th = ["Date", "From Name", "To Name", "Queue", "Agent", "Duration", "Channel", "Full Transcript"]
    tw = [14, 22, 22, 22, 22, 10, 10, 100]
    for col, (h, w) in enumerate(zip(th, tw), 1):
        cell = ws3.cell(row=1, column=col)
        cell.value = h
        cell.font  = Font(name="Calibri", size=11, bold=True, color=WHITE)
        cell.fill  = PatternFill("solid", fgColor=RC_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws3.column_dimensions[get_column_letter(col)].width = w

    for ri, rec in enumerate([r for r in records if r["has_transcript"]], 2):
        try:
            dt = datetime.fromisoformat(rec["start_time"].replace("Z", "+00:00"))
            ds = dt.strftime("%Y-%m-%d")
        except Exception:
            ds = rec["start_time"][:10] if rec["start_time"] else ""
        dm, ds2 = divmod(int(rec["duration_sec"]), 60)
        for col, val in enumerate([
            ds, rec["from_name"], rec["to_name"],
            rec["queue_name"], rec["agent_name"],
            f"{dm}m {ds2}s", rec["channel"], rec["transcript"],
        ], 1):
            cell = ws3.cell(row=ri, column=col)
            cell.value = val
            cell.font  = Font(name="Calibri", size=10, color=DARK)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        ws3.row_dimensions[ri].height = 60

    wb.save(out_path)


# ---------------------------------------------------------------------------
# PDF builder
# ---------------------------------------------------------------------------

def build_pdf(records, customer_name, date_from, date_to, out_path):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable, KeepTogether)
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.pdfgen import canvas as rc_canvas

    RC_BLUE   = HexColor("#0066CC")
    RC_ORANGE = HexColor("#FF6A00")
    DARK      = HexColor("#1A1A1A")
    GREY      = HexColor("#666666")
    BG_BLUE   = HexColor("#E6F1FB")
    BG_ORANGE = HexColor("#FFF3E8")
    BG_GREEN  = HexColor("#E8F8EF")
    BG_GREY   = HexColor("#F5F5F5")
    TX_BLUE   = HexColor("#0B4F8A")
    TX_GREEN  = HexColor("#1B6B35")
    TX_GREY   = HexColor("#444444")
    RULE      = HexColor("#E0E0E0")

    def ps(name, **kw): return ParagraphStyle(name, **kw)

    call_num_s  = ps("CN", fontName="Helvetica",      fontSize=8,  textColor=GREY,   leading=12)
    call_title_s= ps("CT", fontName="Helvetica-Bold", fontSize=13, textColor=DARK,   leading=16, spaceAfter=4)
    meta_s      = ps("ME", fontName="Helvetica",      fontSize=9,  textColor=GREY,   leading=12, spaceAfter=6)
    tr_lbl_s    = ps("TL", fontName="Helvetica-Bold", fontSize=8,  textColor=GREY,   leading=14, spaceBefore=10)
    no_tr_s     = ps("NT", fontName="Helvetica-Oblique", fontSize=9, textColor=GREY, leading=12)
    utt_s       = ps("UT", fontName="Helvetica",      fontSize=9,  textColor=DARK,   leading=13, leftIndent=10, spaceAfter=4)
    sp_styles   = [
        ps("SA", fontName="Helvetica-Bold", fontSize=9, textColor=HexColor("#0B4F8A"), leading=12, spaceBefore=6),
        ps("SB", fontName="Helvetica-Bold", fontSize=9, textColor=HexColor("#8B1A1A"), leading=12, spaceBefore=6),
        ps("SC", fontName="Helvetica-Bold", fontSize=9, textColor=HexColor("#1B6B35"), leading=12, spaceBefore=6),
        ps("SD", fontName="Helvetica-Bold", fontSize=9, textColor=HexColor("#5A189A"), leading=12, spaceBefore=6),
    ]

    class NumCanvas(rc_canvas.Canvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_page_states = []
        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()
        def save(self):
            n = len(self._saved_page_states)
            for state in self._saved_page_states:
                self.__dict__.update(state)
                self._draw_page_number(n)
                super().showPage()
            super().save()
        def _draw_page_number(self, n):
            self.setFont("Helvetica", 8)
            self.setFillColor(HexColor("#AAAAAA"))
            self.drawRightString(letter[0] - 0.6 * inch, 0.4 * inch,
                                 f"Page {self._pageNumber} of {n}")

    doc = SimpleDocTemplate(str(out_path), pagesize=letter,
                            leftMargin=0.75*inch, rightMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    story = []

    # Cover header
    cover_data = [[
        Paragraph("<b>RingCentral RingCX Transcript Report</b>",
                  ps("H1", fontName="Helvetica-Bold", fontSize=18, textColor=HexColor("#FFFFFF"), leading=22)),
        Paragraph(customer_name,
                  ps("H2", fontName="Helvetica-Bold", fontSize=14,
                     textColor=HexColor("#BDD7F5"), leading=18, alignment=TA_RIGHT)),
    ]]
    ct = Table(cover_data, colWidths=[doc.width * 0.6, doc.width * 0.4])
    ct.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), RC_BLUE),
        ("TOPPADDING",    (0, 0), (-1, -1), 14),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 14),
        ("LEFTPADDING",   (0, 0), (0, 0),  18),
        ("RIGHTPADDING",  (-1, 0), (-1, -1), 18),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(ct)
    story.append(Spacer(1, 4))

    sub = Table([[Paragraph(
        f"{date_from} → {date_to} | Generated {datetime.now().strftime('%B %d, %Y')} | Confidential",
        ps("SU", fontName="Helvetica", fontSize=8.5, textColor=HexColor("#555555"),
           alignment=TA_CENTER)
    )]], colWidths=[doc.width])
    sub.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), BG_GREY),
        ("TOPPADDING",    (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(sub)
    story.append(Spacer(1, 16))

    calls_to_show = [r for r in records if r["has_transcript"]] or records

    for idx, row in enumerate(calls_to_show, 1):
        from_name  = row["from_name"]  or row["from_number"] or "Unknown"
        to_name    = row["to_name"]    or row["to_number"]   or "Unknown"
        direction  = row["direction"]
        transcript = row["transcript"]
        queue      = row["queue_name"] or ""
        agent      = row["agent_name"] or ""

        try:
            dt       = datetime.fromisoformat(row["start_time"].replace("Z", "+00:00"))
            date_str = dt.strftime("%B %d, %Y")
            time_str = dt.strftime("%I:%M %p")
        except Exception:
            date_str = row["start_time"][:10] if row["start_time"] else ""
            time_str = ""

        dur_m, dur_s = divmod(int(row["duration_sec"]), 60)
        dir_bg = BG_BLUE   if direction == "Inbound"  else BG_ORANGE
        dir_fg = TX_BLUE   if direction == "Inbound"  else HexColor("#9B4A00")

        block = []
        block.append(Paragraph(f"INTERACTION {idx} OF {len(calls_to_show)}", call_num_s))

        title_text = (f"Inbound from <b>{from_name}</b>"
                      if direction == "Inbound"
                      else f"Outbound to <b>{to_name}</b>")
        block.append(Paragraph(title_text, call_title_s))

        meta_parts = [p for p in [date_str, time_str, f"Duration: {dur_m}m {dur_s}s",
                                  f"Queue: {queue}" if queue else "",
                                  f"Agent: {agent}" if agent else "",
                                  f"Channel: {row['channel']}"] if p]
        block.append(Paragraph(" &nbsp;|&nbsp; ".join(meta_parts), meta_s))

        bt = Table([[
            Paragraph(f"<b>{direction or 'Unknown'}</b>",
                      ps("DB", fontSize=8.5, fontName="Helvetica-Bold",
                         textColor=dir_fg, alignment=TA_CENTER)),
            Paragraph(f"<b>{row['channel']}</b>",
                      ps("CB", fontSize=8.5, fontName="Helvetica-Bold",
                         textColor=TX_GREY, alignment=TA_CENTER)),
            Paragraph("", ps("SP", fontSize=8)),
        ]], colWidths=[1.0*inch, 1.0*inch, doc.width - 2.0*inch])
        bt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, 0), dir_bg),
            ("BACKGROUND", (1, 0), (1, 0), BG_GREY),
            ("TOPPADDING",    (0, 0), (1, 0), 5),
            ("BOTTOMPADDING", (0, 0), (1, 0), 5),
            ("LEFTPADDING",   (0, 0), (1, 0), 10),
            ("RIGHTPADDING",  (0, 0), (1, 0), 10),
        ]))
        block.append(Spacer(1, 4))
        block.append(bt)

        if transcript:
            block.append(Paragraph("FULL TRANSCRIPT", tr_lbl_s))
            unique_sp = list(dict.fromkeys(
                line.split(": ")[0].split("] ")[-1].strip()
                for line in transcript.split("\n") if ": " in line
            ))
            sp_map = {sp: sp_styles[i % len(sp_styles)] for i, sp in enumerate(unique_sp)}

            for line in transcript.split("\n"):
                if not line.strip():
                    continue
                m = re.match(r"^(\[\d+:\d+\])\s+(.+?):\s+(.+)$", line)
                if m:
                    ts2, speaker, text = m.group(1), m.group(2), m.group(3)
                    safe = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                    block.append(Paragraph(
                        f"<b>{speaker}</b> <font size='7' color='#BBBBBB'>{ts2}</font>",
                        sp_map.get(speaker, sp_styles[0])))
                    block.append(Paragraph(safe, utt_s))
                else:
                    safe = line.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                    block.append(Paragraph(safe, utt_s))
        else:
            block.append(Spacer(1, 6))
            block.append(Paragraph(
                "No transcript available — AI transcription may not be enabled for this queue.",
                no_tr_s))

        block.append(Spacer(1, 0.15 * inch))
        block.append(HRFlowable(width="100%", thickness=0.5, color=RULE))
        block.append(Spacer(1, 0.2 * inch))

        story.append(KeepTogether(block[:7]))
        story.extend(block[7:])

    doc.build(story, canvasmaker=NumCanvas)


# ---------------------------------------------------------------------------
# Entrypoint
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print(f"\n RingCentral RingCX Transcript Web App → http://localhost:{port}\n")
    app.run(host="0.0.0.0", port=port, debug=False)
